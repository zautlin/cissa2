#!/usr/bin/env python3
"""
Bloomberg Excel to CSV Extraction Script
Extracts all sheets from an Excel workbook and converts to CSV format with cleaning.

Usage:
    python3 01_extract_excel_to_csv.py <input_excel_file> <output_directory>

Example:
    python3 01_extract_excel_to_csv.py "./raw-data/Bloomberg Download data.xlsx" "./extracted-worksheets"
"""

import sys
import os
import csv
from pathlib import Path
from openpyxl import load_workbook


def extract_and_clean_excel(excel_file, output_dir):
    """
    Extract all sheets from Excel workbook and save as cleaned CSV files.
    
    Process:
    1. Load Excel workbook
    2. For each sheet:
       - Identify proper header row (skipping Excel column headers, blanks, errors)
       - Remove unnecessary rows (blanks, errors, instructions)
       - Save as CSV with proper headers
    """
    
    # Validate input file
    excel_path = Path(excel_file)
    if not excel_path.exists():
        print(f"❌ Error: Excel file not found: {excel_file}")
        sys.exit(1)
    
    # Create output directory
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    print("=" * 80)
    print("📊 BLOOMBERG EXCEL TO CSV EXTRACTION")
    print("=" * 80)
    print(f"Input:  {excel_path}")
    print(f"Output: {output_path}")
    print()
    
    # Load workbook
    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        sys.exit(1)
    
    print(f"Found {len(wb.sheetnames)} sheets")
    print()
    
    stats = {
        'sheets_processed': 0,
        'sheets_empty': 0,
        'total_rows_extracted': 0,
        'total_rows_removed': 0,
        'csv_files_created': 0
    }
    
    # Process each sheet
    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        
        print(f"[{sheet_idx}/{len(wb.sheetnames)}] Processing: {sheet_name}")
        
        # Get all rows
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(row)
        
        if not rows:
            print(f"    ⊘ Sheet is empty, skipping")
            stats['sheets_empty'] += 1
            continue
        
        stats['sheets_processed'] += 1
        
        # Find the actual header row
        # Pattern: Row 0 usually has A,B,C,D..., row 13 has actual headers
        header_row_idx = None
        for idx, row in enumerate(rows):
            # Check if this looks like a header row (starts with Num or contains expected headers)
            if row and len(row) > 0:
                row_str = str(row[0]).lower() if row[0] else ""
                # Common header patterns
                if any(pattern in row_str for pattern in ['num', 'ticker', 'name']):
                    header_row_idx = idx
                    break
        
        # If no header found, skip sheet
        if header_row_idx is None:
            print(f"    ⊘ No header row found, skipping")
            stats['sheets_empty'] += 1
            continue
        
        # Extract headers
        headers = [str(h) if h else "" for h in rows[header_row_idx]]
        
        # Extract data rows (skip blank/error rows)
        data_rows = []
        rows_removed = 0
        
        for idx in range(header_row_idx + 1, len(rows)):
            row = rows[idx]
            
            # Skip if row is completely empty
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                rows_removed += 1
                continue
            
            # Skip if row contains only #REF! errors
            if all(str(cell) == "#REF!" for cell in row if cell is not None):
                rows_removed += 1
                continue
            
            # Skip if first column suggests it's not data (common patterns)
            first_cell = str(row[0]).lower() if row[0] else ""
            if any(skip in first_cell for skip in ['replace', 'blank', 'bloomberg']):
                rows_removed += 1
                continue
            
            # This is a valid data row
            data_rows.append(row)
        
        # Create CSV filename from sheet name
        csv_filename = f"{sheet_name}.csv"
        csv_path = output_path / csv_filename
        
        # Write CSV file
        try:
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write headers
                writer.writerow(headers)
                
                # Write data rows
                for row in data_rows:
                    # Convert None to empty string for CSV
                    row_clean = [str(cell) if cell is not None else "" for cell in row]
                    writer.writerow(row_clean)
            
            print(f"    ✓ Created: {csv_filename}")
            print(f"      Headers: {len(headers)} columns")
            print(f"      Data rows: {len(data_rows)} (removed {rows_removed} unnecessary rows)")
            
            stats['csv_files_created'] += 1
            stats['total_rows_extracted'] += len(data_rows)
            stats['total_rows_removed'] += rows_removed
            
        except Exception as e:
            print(f"    ❌ Error writing CSV: {e}")
    
    # Print summary
    print()
    print("=" * 80)
    print("✅ EXTRACTION COMPLETE")
    print("=" * 80)
    print(f"Sheets processed:        {stats['sheets_processed']}")
    print(f"Sheets empty/skipped:    {stats['sheets_empty']}")
    print(f"CSV files created:       {stats['csv_files_created']}")
    print(f"Total data rows:         {stats['total_rows_extracted']}")
    print(f"Total rows removed:      {stats['total_rows_removed']}")
    print(f"Output directory:        {output_path}")
    print("=" * 80)
    print()


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 01_extract_excel_to_csv.py <input_excel> <output_dir>")
        print()
        print("Example:")
        print('  python3 01_extract_excel_to_csv.py "./raw-data/Bloomberg Download data.xlsx" "./extracted-worksheets"')
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_dir = sys.argv[2]
    
    extract_and_clean_excel(excel_file, output_dir)
